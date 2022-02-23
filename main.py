from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import constants
import time
import openpyxl
import xlrd
import xlwt

book = openpyxl.Workbook()
sheet = book.active
row = 1
sheet.cell(row=row, column=1).value = 'Адрес'
sheet.cell(row=row, column=2).value = 'Цена'
sheet.cell(row=row, column=3).value = 'Ссылка'
row += 1

s = Service(ChromeDriverManager().install())  # это установка драйвера
options = webdriver.ChromeOptions()
options.add_argument('headless')
options.add_argument('window-size=1366x768')
driver = webdriver.Chrome(service=s)

try:
    driver.get(constants.URL)  # ссылка на сайт который парсим
    time.sleep(1)
    data = []
    while True:

        pricelist = driver.find_elements(By.XPATH, constants.pricelist)
        adress = driver.find_elements(By.XPATH, constants.adress)
        href = driver.find_elements(By.XPATH, constants.href)

        try:
            pricelist_xl = driver.find_elements(By.XPATH, constants.pricelist_xl)
            adress_xl = driver.find_elements(By.XPATH, constants.adress_xl)
            href_xl = driver.find_elements(By.XPATH, constants.href_xl)

            pricelist += pricelist_xl
            adress += adress_xl
            href += href_xl
        except:
            pass
        # print(len(pricelist))
        for i in range(len(pricelist)):
            data.append([adress[i].text, pricelist[i].text, href[i].get_attribute('href')])
            # sheet.cell(row=row, column=1).value = adress[i].text
            # sheet.cell(row=row, column=2).value = pricelist[i].text
            # sheet.cell(row=row, column=3).value = href[i].get_attribute('href')
            # print(adress[i].text, pricelist[i].text, href[i].get_attribute('href'))
        time.sleep(1)
        try:
            action = ActionChains(driver)
            element = driver.find_element(By.XPATH, constants.NEXTBUTTON)
            action.move_to_element(element).perform()
            element.click()
            print(driver.current_url)
            time.sleep(5)
        except:
            break
    data.sort(key = lambda x:x[1])
    for i in range(len(data)):
        sheet.cell(row=i+2, column=1).value = data[i][0]
        sheet.cell(row=i+2, column=2).value = data[i][1]
        sheet.cell(row=i+2, column=3).value = data[i][2]

    book.save("C://Users/k1r9n/Desktop/parser.xlsx")
    book.close()

except Exception as ex:
    print(ex)

finally:
    driver.close()
    driver.quit()
